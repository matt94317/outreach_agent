#!/usr/bin/env python3
"""
OutreachAI — Local Email Outreach Agent
========================================
Reads your client spreadsheet, personalises emails with Claude AI,
opens a pre-filled Gmail compose window in your browser for review,
waits for your confirmation in the terminal, then marks the
spreadsheet as Sent or Skipped.

Flow per client:
  1. AI generates personalised email
  2. Terminal shows a preview
  3. Press [o] → Gmail compose opens in your browser, pre-filled
  4. Review / edit inside Gmail, then send from there
  5. Back in terminal: [y] confirm sent · [n] didn't send / skip · [e] regenerate

Usage:
    python outreach_agent.py
"""

import csv
import json
import os
import re
import time
import webbrowser
import urllib.parse
import textwrap
import sys
from datetime import datetime
from pathlib import Path

# Load .env file if present
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# ── Optional Excel support ────────────────────────────────────────────────────
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ── Optional Anthropic SDK (falls back to raw HTTP) ───────────────────────────
try:
    import anthropic
    USE_SDK = True
except ImportError:
    import urllib.request
    USE_SDK = False

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  Edit this section before running
# ══════════════════════════════════════════════════════════════════════════════

CONFIG = {
    # ── Spreadsheet ───────────────────────────────────────────────────────────
    "spreadsheet_path": "HOOKD_Master_Outreach_v4.xlsx",   # CSV or .xlsx file
    # Column names in your spreadsheet (case-insensitive, partial match ok)
    "col_company":          "Company Name",
    "col_contact_name":     "Contact Name",
    "col_role":             "Role",
    "col_email":            "Email",            # matches "Email(s)"
    "col_linkedin":         "LinkedIn",         # matches "Linke..."
    "col_date_contacted":   "Date Last Contacted",
    "col_website":      "Website",
    "col_status":           "Status",           # Will be added/updated automatically
    "col_action_required":  "Action Required",
    "col_language":         "Language",         # Optional: "EN" → English template, anything else → Chinese

    # ── Your details ─────────────────────────────────────────────────────────
    "sender_name":     "Renee Yu",
    "sender_title":    "Hookd 共同創辦人",
    "sender_company":  "Hookd",
    "sender_website":  "https://hookdugc.framer.website/",
    "booking_link":    "https://calendar.app.google/2Gh1kTgRiH88sn5J6",   # Google Meet / Calendly
    "booking_label":   "Book a free 30-min consultation",

    # ── Gmail ─────────────────────────────────────────────────────────────────
    # Just your Gmail address — used to pre-fill the From field in compose URLs.
    # No password needed. Gmail opens in your browser and you send from there.
    "gmail_address":   "renee@hookdugc.com",

    # ── Anthropic API key ─────────────────────────────────────────────────────
    # Set ANTHROPIC_API_KEY in your .env file (never hardcode secrets here)
    "anthropic_api_key": os.environ.get("ANTHROPIC_API_KEY", ""),

    # ── Email template ────────────────────────────────────────────────────────
    # Use {{company_name}}, {{website}}, {{ai_insight}} as placeholders.
    "email_subject_template": "Helping {{company_name}} grow with Digital Marketing",

    "email_body_template": """
團隊您好，
我是來自 HOOKd 的 Renee。
我們已協助全球多個品牌透過 UGC 行銷，提升至百萬的月下載量
我們一直非常佩服 [公司名稱] 在 [該公司的核心價值/產品亮點，例如：思維導圖領域的深耕／建構女性安全社交空間上的貢獻]，特別是 [具體特色描述，例如：將複雜的邏輯轉化為優雅視覺圖表／透過直播與語音房間讓社群跨越地理限制]，這在當前市場中極具價值。
對於 [公司名稱] 而言，「[核心吸引新用戶的素材方向]」 是最吸引新用戶的素材。
HOOKd 專門製作高轉化率的 UGC（用戶原創內容），能協助您：

[內容方向一，例如：沉浸式靈感拆解]： [具體說明，例如：透過創作者演示縮時過程，展示產品絲滑的操作感與多樣化功能]
[內容方向二，例如：場景化知識整理]： [具體說明，例如：製作情境短影音，降低工具使用門檻，引發社群儲存與下載]
[內容方向三，例如：高質感品牌推廣]： 針對 IG Reels/TikTok 產出具有美學厚度的原生素材，避開生硬的軟體操作教學，以「[切入角度]」切入，優化獲客成本（CPI）

希望能撥冗 30 分鐘與您交流，分享我們如何協助 [公司名稱] 在社群創造更高品質的增長。

""",

    # ── English email template ────────────────────────────────────────────────
    # Used when a row's Language column is "EN" (case-insensitive).
    "email_subject_template_en": "Helping {{company_name}} grow with UGC Marketing",

    "email_body_template_en": """
Hi [Team / Contact Name],

My name is Renee, co-founder of HOOKd✨.

We've helped brands worldwide scale to **millions of monthly downloads** through **UGC (User-Generated Content) marketing**.

We've been genuinely impressed by [Company Name]'s work in [core value/product highlight — e.g. the way you simplify complex workflows / your commitment to building meaningful community experiences]. That kind of [specific differentiator] is exactly what resonates with today's audiences.

For [Company Name], we believe **core content hook** is the strongest angle to attract new users at scale.

**HOOKd specialises in high-converting UGC** and can help you with:

    • **[Content direction 1 — e.g. Immersive product storytelling]:** [Brief description — e.g. Creators showcase your product's core moments, driving curiosity and installs]
    • **[Content direction 2 — e.g. Scenario-based use cases]:** [Brief description — e.g. Short-form videos that lower the barrier to entry and spark saves & shares]
    • **[Content direction 3 — e.g. Native social creatives]:** Platform-native content for TikTok & Instagram Reels — leading with [angle], optimising your **CPI** without the hard-sell feel

Would love to grab **30 minutes** to share how we can help [Company Name] build higher-quality growth on social.

""",

    # ── Behaviour ─────────────────────────────────────────────────────────────
    "delay_between_sends": 2,     # seconds to wait between sent emails
    "skip_already_sent":   True,  # skip rows already marked "Sent"
}

# ══════════════════════════════════════════════════════════════════════════════
#  COLOURS & TERMINAL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

class C:
    RESET  = "\033[0m"
    BOLD   = "\033[1m"
    DIM    = "\033[2m"
    GREEN  = "\033[92m"
    YELLOW = "\033[93m"
    RED    = "\033[91m"
    CYAN   = "\033[96m"
    WHITE  = "\033[97m"
    BLUE   = "\033[94m"
    MAGENTA= "\033[95m"

def banner():
    print(f"""
{C.CYAN}{C.BOLD}
  ╔═══════════════════════════════════════╗
  ║        OutreachAI  —  v2.0            ║
  ║   Local Email Outreach Agent 🤖       ║
  ║   Gmail Compose Mode                  ║
  ╚═══════════════════════════════════════╝
{C.RESET}""")

def section(title):
    print(f"\n{C.BOLD}{C.WHITE}{'─'*50}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}  {title}{C.RESET}")
    print(f"{C.BOLD}{C.WHITE}{'─'*50}{C.RESET}")

def info(msg):    print(f"  {C.BLUE}ℹ{C.RESET}  {msg}")
def success(msg): print(f"  {C.GREEN}✓{C.RESET}  {C.GREEN}{msg}{C.RESET}")
def warn(msg):    print(f"  {C.YELLOW}⚠{C.RESET}  {C.YELLOW}{msg}{C.RESET}")
def error(msg):   print(f"  {C.RED}✗{C.RESET}  {C.RED}{msg}{C.RESET}")
def dim(msg):     print(f"  {C.DIM}{msg}{C.RESET}")

def divider(): print(f"  {C.DIM}{'·'*46}{C.RESET}")

# ══════════════════════════════════════════════════════════════════════════════
#  SPREADSHEET  READ / WRITE
# ══════════════════════════════════════════════════════════════════════════════

def find_col(headers, keyword):
    """Find a column header that contains the keyword (case-insensitive)."""
    kw = keyword.lower()
    for h in headers:
        if kw in h.lower():
            return h
    return None

def _sheet_language(sheet_name):
    """Infer language from sheet tab name. Returns 'EN' or 'ZH'."""
    name = sheet_name.strip().upper()
    if "EN" in name or "ENGLISH" in name:
        return "EN"
    return "ZH"

def _parse_sheet(ws):
    """Parse one openpyxl worksheet into (headers, rows) using the same format as CSV DictReader."""
    raw = list(ws.values)
    if not raw:
        return [], []
    headers = [str(h) if h is not None else "" for h in raw[0]]
    rows = []
    for row in raw[1:]:
        rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
    return headers, rows

def load_clients(path):
    path = Path(path)
    if not path.exists():
        error(f"Spreadsheet not found: {path}")
        error("Please set 'spreadsheet_path' in CONFIG and make sure the file exists.")
        sys.exit(1)

    ext = path.suffix.lower()
    # rows and headers are dicts keyed by sheet name for xlsx,
    # or use the sentinel key "_csv" for csv files.
    all_rows    = {}   # {sheet_name: [row_dict, ...]}
    all_headers = {}   # {sheet_name: [col_name, ...]}

    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            hdrs = reader.fieldnames or []
            rws  = list(reader)
        all_headers["_csv"] = hdrs
        all_rows["_csv"]    = rws

    elif ext in (".xlsx", ".xls"):
        if not EXCEL_SUPPORT:
            error("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            hdrs, rws = _parse_sheet(ws)
            if hdrs:
                all_headers[ws.title] = hdrs
                all_rows[ws.title]    = rws
        if not all_headers:
            error("Spreadsheet appears to be empty.")
            sys.exit(1)
    else:
        error(f"Unsupported file type: {ext}. Use .csv or .xlsx")
        sys.exit(1)

    cfg = CONFIG
    clients   = []
    col_status = None   # resolved per sheet (assumed same name across sheets)

    for sheet_name, headers in all_headers.items():
        rows     = all_rows[sheet_name]
        language = _sheet_language(sheet_name) if ext != ".csv" else "ZH"

        col_company        = find_col(headers, cfg["col_company"]) or (headers[0] if headers else "company")
        col_contact_name   = find_col(headers, cfg["col_contact_name"])
        col_role           = find_col(headers, cfg["col_role"])
        col_email          = find_col(headers, cfg["col_email"])
        col_linkedin       = find_col(headers, cfg["col_linkedin"])
        col_date_contacted = find_col(headers, cfg["col_date_contacted"])
        col_description    = find_col(headers, cfg["col_website"])
        col_status_sheet   = find_col(headers, cfg["col_status"])
        col_action         = find_col(headers, cfg["col_action_required"])
        col_language_col   = find_col(headers, cfg["col_language"])
        if col_status_sheet:
            col_status = col_status_sheet

        info(f"Sheet '{sheet_name}' → default language: {language}, {len(rows)} data rows")

        for i, row in enumerate(rows):
            company      = row.get(col_company, "").strip()
            email        = row.get(col_email, "").strip()          if col_email          else ""
            contact_name = row.get(col_contact_name, "").strip()   if col_contact_name   else ""
            role         = row.get(col_role, "").strip()           if col_role           else ""
            linkedin     = row.get(col_linkedin, "").strip()       if col_linkedin        else ""
            date_contacted = row.get(col_date_contacted, "").strip() if col_date_contacted else ""
            description  = row.get(col_description, "").strip()   if col_description    else ""
            status       = row.get(col_status_sheet, "").strip()  if col_status_sheet   else ""
            action       = row.get(col_action, "").strip()         if col_action         else ""
            # Per-row Language column overrides the sheet-level default
            lang_cell    = row.get(col_language_col, "").strip().upper() if col_language_col else ""
            row_language = "EN" if "EN" in lang_cell else ("ZH" if lang_cell else language)
            if not company and not email:
                continue
            clients.append({
                "_row_index":  i,
                "_sheet":      sheet_name,
                "_raw":        row,
                "company":     company,
                "contact_name": contact_name,
                "role":        role,
                "email":       email,
                "linkedin":    linkedin,
                "date_contacted": date_contacted,
                "website":     description,
                "status":      status,
                "action":      action,
                "language":    row_language,
            })

    return clients, all_headers, all_rows, None, None, None, col_status, path, ext

def save_clients(clients, all_headers, all_rows, col_status, path, ext):
    """Write status updates back to the spreadsheet."""
    col_status_name = col_status or "status"

    # Push updated statuses back into the per-sheet row dicts
    for client in clients:
        sheet = client.get("_sheet", "_csv")
        idx   = client["_row_index"]
        sheet_rows = all_rows.get(sheet, [])
        if idx < len(sheet_rows):
            sheet_rows[idx][col_status_name] = client["status"]

    if ext == ".csv":
        headers = list(all_headers.get("_csv", []))
        rows    = all_rows.get("_csv", [])
        if col_status_name not in headers:
            headers.append(col_status_name)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                if col_status_name not in row:
                    row[col_status_name] = ""
                writer.writerow(row)
    else:
        if not EXCEL_SUPPORT:
            warn("openpyxl not installed, cannot save .xlsx. Install with: pip install openpyxl")
            return
        wb = openpyxl.Workbook()
        for i, (sheet_name, headers) in enumerate(all_headers.items()):
            headers = list(headers)
            if col_status_name not in headers:
                headers.append(col_status_name)
            rows = all_rows.get(sheet_name, [])
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = sheet_name
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
        wb.save(path)

# ══════════════════════════════════════════════════════════════════════════════
#  AI EMAIL GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def generate_email(client):
    """Call Claude to personalise the email for a specific client."""
    cfg = CONFIG
    api_key = cfg["anthropic_api_key"]
    is_english = client.get("language", "").strip().upper() == "EN"
    template    = cfg["email_body_template_en"]    if is_english else cfg["email_body_template"]
    subject_tpl = cfg["email_subject_template_en"] if is_english else cfg["email_subject_template"]

    website_ctx = f"Their website is: {client['website']}." if client["website"] else "No website provided."

    prompt = f"""You are an expert digital marketing outreach specialist writing a personalised cold email.

Client details:
- Company name: {client['company'] or 'Unknown'}
- Email: {client['email'] or 'Unknown'}
- {website_ctx}

Email body template to personalise:
---
{template}
---

Subject template:
{subject_tpl}

Instructions:
- Personalise the subject and body specifically for this company
- Replace {{{{company_name}}}} with the actual company name
- Replace {{{{website}}}} with their website if available
- Replace {{{{ai_insight}}}} with 2-3 specific, compelling sentences about a digital marketing opportunity relevant to THIS type of business (based on their name/website)
- Keep {{{{sender_name}}}}, {{{{sender_title}}}}, {{{{sender_company}}}}, {{{{sender_website}}}}, {{{{booking_link}}}} as-is — these will be filled in separately
- Keep the tone professional yet warm and conversational
- Do NOT be generic — make it feel like you genuinely researched them
- Keep the body under 200 words
- Wrap key phrases and important keywords in 【】 brackets (e.g. 【UGC行銷】 or 【millions of downloads】) — use these for company achievements, service names, key benefits, metrics, and calls to action

Return ONLY valid JSON, nothing else:
{{"subject": "...", "body": "..."}}"""

    if USE_SDK:
        client_sdk = anthropic.Anthropic(api_key=api_key)
        message = client_sdk.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        text = message.content[0].text
    else:
        import urllib.request, json as jsonlib
        payload = jsonlib.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 1024,
            "messages": [{"role": "user", "content": prompt}]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            }
        )
        with urllib.request.urlopen(req) as resp:
            data = jsonlib.loads(resp.read())
        text = data["content"][0]["text"]

    # Parse JSON response
    clean = text.strip()
    if clean.startswith("```"):
        clean = clean.split("```")[1]
        if clean.startswith("json"):
            clean = clean[4:]
    clean = clean.strip()
    result = json.loads(clean)
    return result["subject"], result["body"]

def fill_sender_placeholders(text):
    """Replace sender-side placeholders with config values."""
    cfg = CONFIG
    replacements = {
        "{{sender_name}}":    cfg["sender_name"],
        "{{sender_title}}":   cfg["sender_title"],
        "{{sender_company}}": cfg["sender_company"],
        "{{sender_website}}": cfg["sender_website"],
        "{{booking_link}}":   cfg["booking_link"],
        "{{booking_label}}":  cfg["booking_label"],
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text

def build_signature():
    cfg = CONFIG
    sig = f"""
--
📅 {cfg['booking_label']}
   30 minutes · Google Meet
   Book here: {cfg['booking_link']}

{cfg['sender_name']}
{cfg['sender_title']} | {cfg['sender_company']}
{cfg['sender_website']}
"""
    return sig

# ══════════════════════════════════════════════════════════════════════════════
#  TERMINAL REVIEW UI
# ══════════════════════════════════════════════════════════════════════════════

def render_highlight(text):
    """Render 【keyword】 brackets in ANSI yellow for terminal display."""
    return re.sub(r'【(.+?)】', lambda m: f"{C.YELLOW}【{m.group(1)}】{C.RESET}{C.WHITE}", text)

def display_email_preview(client, subject, body, index, total):
    section(f"Client {index}/{total}  —  {client['company'] or client['email']}")

    print(f"\n  {C.DIM}Company :{C.RESET} {C.WHITE}{client['company']}{C.RESET}")
    print(f"  {C.DIM}Email   :{C.RESET} {C.CYAN}{client['email']}{C.RESET}")
    if client["website"]:
        print(f"  {C.DIM}Website :{C.RESET} {C.BLUE}{client['website']}{C.RESET}")

    divider()
    print(f"\n  {C.BOLD}Subject:{C.RESET} {C.YELLOW}{subject}{C.RESET}\n")

    # Word-wrap body for terminal display, highlighting 【keyword】 brackets
    for line in body.splitlines():
        if line.strip():
            wrapped = textwrap.fill(line, width=60, initial_indent="  ", subsequent_indent="  ")
            print(f"{C.WHITE}{render_highlight(wrapped)}{C.RESET}")
        else:
            print()


def prompt_approval(client):
    """
    New two-step Gmail flow:
      [o]  Open Gmail compose in browser (pre-filled)
      [y]  Confirm you sent it from Gmail
      [n]  Skip this client
      [e]  Edit subject/body then re-preview
      [q]  Quit and save progress
    """
    while True:
        print(f"\n  {C.BOLD}What would you like to do?{C.RESET}")
        print(f"  {C.CYAN}[o]{C.RESET} Open in Gmail  {C.DIM}(pre-fills To, Subject & Body){C.RESET}")
        print(f"  {C.GREEN}[y]{C.RESET} Confirm sent   {C.DIM}(after you clicked Send in Gmail){C.RESET}")
        print(f"  {C.RED}[n]{C.RESET} Skip")
        print(f"  {C.YELLOW}[e]{C.RESET} Edit subject or body")
        print(f"  {C.MAGENTA}[q]{C.RESET} Quit and save progress")

        choice = input(f"\n  {C.BOLD}Your choice (o/y/n/e/q): {C.RESET}").strip().lower()

        if choice == "o":
            return "open", None, None
        elif choice in ("y", "yes"):
            return "confirm_sent", None, None
        elif choice in ("n", "no", "s", "skip"):
            return "skip", None, None
        elif choice == "e":
            return "edit", None, None
        elif choice == "q":
            return "quit", None, None
        else:
            warn("Please type o, y, n, e, or q")

def edit_email(subject, body):
    """Allow inline editing of subject and body."""
    print(f"\n  {C.YELLOW}Edit subject{C.RESET} (press Enter to keep current):")
    print(f"  Current: {subject}")
    new_subject = input(f"  New subject: ").strip()
    if not new_subject:
        new_subject = subject

    print(f"\n  {C.YELLOW}Edit body{C.RESET}")
    print(f"  (Tip: open your editor, paste new content, then paste it here)")
    print(f"  Type your new body below, then type {C.BOLD}END{C.RESET} on a new line (or press Enter to keep current):\n")

    lines = []
    while True:
        line = input()
        if line.strip().upper() == "END":
            break
        lines.append(line)

    new_body = "\n".join(lines).strip() if lines else body
    return new_subject, new_body

# ══════════════════════════════════════════════════════════════════════════════
#  GMAIL COMPOSE  —  Opens pre-filled compose window in browser
# ══════════════════════════════════════════════════════════════════════════════

def open_gmail_compose(to_email, subject, body):
    """
    Opens Gmail compose in the default browser with To, Subject, and Body
    pre-filled. The user reviews and clicks Send inside Gmail.
    """
    params = urllib.parse.urlencode({
        "to":   to_email,
        "su":   subject,
        "body": body,
    })
    url = f"https://mail.google.com/mail/?view=cm&fs=1&{params}"
    webbrowser.open(url)

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN AGENT LOOP
# ══════════════════════════════════════════════════════════════════════════════

def main():
    banner()

    # ── Load spreadsheet ──────────────────────────────────────────────────────
    section("Loading client list")
    (clients, headers, rows,
     col_company, col_email, _,
     col_status, path, ext) = load_clients(CONFIG["spreadsheet_path"])

    total = len(clients)
    info(f"Loaded {C.BOLD}{total}{C.RESET} clients from {C.CYAN}{path.name}{C.RESET}")

    # Skip any row whose status starts with "sent" (case-insensitive)
    to_process = []
    skipped_count = 0
    for c in clients:
        if c["status"].lower().startswith("sent"):
            skipped_count += 1
        else:
            to_process.append(c)

    if skipped_count:
        info(f"Skipping {skipped_count} already-sent clients")

    if not to_process:
        success("All clients have already been contacted. Nothing to do!")
        return

    info(f"{C.BOLD}{len(to_process)}{C.RESET} clients to process\n")

    # ── Validate config ───────────────────────────────────────────────────────
    if CONFIG["anthropic_api_key"].startswith("sk-ant-..."):
        error("Please set your Anthropic API key in CONFIG['anthropic_api_key']")
        sys.exit(1)
    if CONFIG["gmail_address"] == "you@gmail.com":
        error("Please set your Gmail address in CONFIG['gmail_address']")
        sys.exit(1)
    # ── Stats tracking ────────────────────────────────────────────────────────
    stats = {"sent": 0, "skipped": 0, "errors": 0}

    # ── Process each client ───────────────────────────────────────────────────
    for i, client in enumerate(to_process, 1):
        try:
            # Generate email
            info(f"Generating email for {C.BOLD}{client['company'] or client['email']}{C.RESET}...")
            subject, body = generate_email(client)
            subject = fill_sender_placeholders(subject)
            body    = fill_sender_placeholders(body)

        except Exception as e:
            warn(f"AI generation failed for {client['company']}: {e}")
            warn("Using template fallback")
            subject = fill_sender_placeholders(
                CONFIG["email_subject_template"].replace("{{company_name}}", client["company"])
            )
            body = fill_sender_placeholders(
                CONFIG["email_body_template"]
                    .replace("{{company_name}}", client["company"])
                    .replace("{{website}}", client["website"])
                    .replace("{{ai_insight}}", "We see great potential to help grow your digital presence.")
            )

        # Review loop
        gmail_opened = False
        while True:
            display_email_preview(client, subject, body, i, len(to_process))
            action, _, _ = prompt_approval(client)

            if action == "edit":
                subject, body = edit_email(subject, body)
                gmail_opened = False   # reset so they can re-open with edits
                continue

            elif action == "open":
                if not client["email"]:
                    warn(f"No email address for {client['company']} — cannot open Gmail")
                    continue
                info(f"Opening Gmail compose for {C.CYAN}{client['email']}{C.RESET}...")
                open_gmail_compose(client["email"], subject, body)
                gmail_opened = True
                info("Gmail opened in your browser.")
                info(f"Review the email, then click {C.BOLD}Send{C.RESET} inside Gmail.")
                info(f"Come back here and press {C.GREEN}[y]{C.RESET} to confirm, or {C.RED}[n]{C.RESET} to skip.")
                continue

            elif action == "confirm_sent":
                if not gmail_opened:
                    warn("You haven't opened Gmail yet — press [o] first to open and send the email.")
                    continue
                client["status"] = f"Sent {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                stats["sent"] += 1
                success(f"Marked as sent · {client['email']}")
                time.sleep(CONFIG["delay_between_sends"])
                break

            elif action == "skip":
                client["status"] = "Skipped"
                stats["skipped"] += 1
                info("Skipped.")
                break

            elif action == "quit":
                info("Saving progress and quitting...")
                save_clients(clients, headers, rows, col_status, path, ext)
                success(f"Progress saved to {path.name}")
                print_summary(stats, len(to_process))
                return

        # Auto-save after every client
        save_clients(clients, headers, rows, col_status, path, ext)

    # ── Done ──────────────────────────────────────────────────────────────────
    section("All done!")
    save_clients(clients, headers, rows, col_status, path, ext)
    success(f"Spreadsheet updated: {path.name}")
    print_summary(stats, len(to_process))

def print_summary(stats, total):
    print(f"""
  {C.BOLD}Summary{C.RESET}
  {C.GREEN}✓ Sent   : {stats['sent']}{C.RESET}
  {C.DIM}— Skipped: {stats['skipped']}{C.RESET}
  {C.RED}✗ Errors : {stats['errors']}{C.RESET}
  {C.DIM}─ Total  : {total}{C.RESET}
""")

if __name__ == "__main__":
    main()
